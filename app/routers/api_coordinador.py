"""
POSTULANDO - API Endpoints para Gestión del Coordinador
app/routers/api_coordinador.py

Endpoints para:
- CRUD de Postulantes
- CRUD de Aulas
- CRUD de Profesores
- Generación de Listados
- Carga masiva (Excel/CSV)
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, func, or_
from typing import Optional, List
from datetime import datetime
import io
import csv

from app.database import get_db
from app.services.auth_admin import obtener_usuario_actual
from app.models import Postulante, Aula, Profesor

router = APIRouter(prefix="/admin/api", tags=["API Coordinador"])


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def obtener_proceso_actual():
    """Obtiene el proceso de admisión actual"""
    now = datetime.now()
    semestre = "1" if now.month <= 6 else "2"
    return f"{now.year}-{semestre}"


def verificar_permisos_coordinador(usuario: dict):
    """Verifica que el usuario tenga permisos de coordinador"""
    if usuario.get('rol') not in ['COORDINADOR', 'DIRECTOR', 'ADMIN']:
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")


# ============================================================
# ESTADÍSTICAS GENERALES
# ============================================================

@router.get("/stats")
async def obtener_estadisticas(
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Obtiene estadísticas generales del proceso"""
    proceso = obtener_proceso_actual()
    
    result = db.execute(text("""
        SELECT 
            (SELECT COUNT(*) FROM postulantes WHERE proceso_admision = :proceso AND activo = true) as total_postulantes,
            (SELECT COUNT(*) FROM aulas WHERE activo = true) as total_aulas,
            (SELECT COUNT(*) FROM profesores WHERE activo = true) as total_profesores,
            (SELECT COUNT(DISTINCT ae.postulante_id) FROM asignaciones_examen ae 
             JOIN postulantes p ON p.id = ae.postulante_id 
             WHERE p.proceso_admision = :proceso AND p.activo = true) as postulantes_asignados,
            (SELECT COUNT(*) FROM hojas_respuestas WHERE proceso_admision = :proceso) as hojas_procesadas
    """), {"proceso": proceso}).fetchone()
    
    return {
        "total_postulantes": result.total_postulantes or 0,
        "total_aulas": result.total_aulas or 0,
        "total_profesores": result.total_profesores or 0,
        "postulantes_asignados": result.postulantes_asignados or 0,
        "hojas_procesadas": result.hojas_procesadas or 0,
        "proceso_actual": proceso
    }


# ============================================================
# CRUD POSTULANTES
# ============================================================

@router.get("/postulantes")
async def listar_postulantes(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    dni: Optional[str] = None,
    apellidos: Optional[str] = None,
    programa: Optional[str] = None,
    turno: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Lista postulantes con filtros y paginación"""
    proceso = obtener_proceso_actual()
    
    # Construir query base con SQL para obtener aula asignada
    base_query = """
        SELECT 
            p.id, p.dni, p.nombres, p.apellido_paterno, p.apellido_materno,
            p.programa_educativo, p.examen_rendido, p.activo,
            a.codigo as aula_codigo
        FROM postulantes p
        LEFT JOIN asignaciones_examen ae ON ae.postulante_id = p.id
        LEFT JOIN aulas a ON a.id = ae.aula_id
        WHERE p.proceso_admision = :proceso AND p.activo = true
    """
    
    count_query = """
        SELECT COUNT(*) 
        FROM postulantes p
        WHERE p.proceso_admision = :proceso AND p.activo = true
    """
    
    params = {"proceso": proceso}
    
    # Aplicar filtros
    if dni:
        base_query += " AND p.dni ILIKE :dni"
        count_query += " AND p.dni ILIKE :dni"
        params["dni"] = f"%{dni}%"
    if apellidos:
        base_query += " AND (p.apellido_paterno ILIKE :apellidos OR p.apellido_materno ILIKE :apellidos)"
        count_query += " AND (p.apellido_paterno ILIKE :apellidos OR p.apellido_materno ILIKE :apellidos)"
        params["apellidos"] = f"%{apellidos}%"
    if programa:
        base_query += " AND p.programa_educativo = :programa"
        count_query += " AND p.programa_educativo = :programa"
        params["programa"] = programa
    
    # Contar total
    total = db.execute(text(count_query), params).scalar() or 0
    
    # Paginar
    offset = (page - 1) * per_page
    base_query += " ORDER BY p.apellido_paterno, p.apellido_materno, p.nombres"
    base_query += f" LIMIT {per_page} OFFSET {offset}"
    
    rows = db.execute(text(base_query), params).fetchall()
    
    postulantes_data = []
    for row in rows:
        postulantes_data.append({
            "id": row.id,
            "dni": row.dni,
            "nombres": row.nombres,
            "apellido_paterno": row.apellido_paterno,
            "apellido_materno": row.apellido_materno,
            "programa_educativo": row.programa_educativo,
            "turno": "MAÑANA",  # Por defecto ya que no hay campo turno
            "aula_codigo": row.aula_codigo,
            "examen_rendido": row.examen_rendido,
            "activo": row.activo
        })
    
    return {
        "postulantes": postulantes_data,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }


@router.get("/postulantes/{postulante_id}")
async def obtener_postulante(
    postulante_id: int,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Obtiene un postulante por ID"""
    postulante = db.query(Postulante).filter(Postulante.id == postulante_id).first()
    
    if not postulante:
        raise HTTPException(status_code=404, detail="Postulante no encontrado")
    
    return {
        "id": postulante.id,
        "dni": postulante.dni,
        "nombres": postulante.nombres,
        "apellido_paterno": postulante.apellido_paterno,
        "apellido_materno": postulante.apellido_materno,
        "programa_educativo": postulante.programa_educativo,
        "turno": getattr(postulante, 'turno', 'MAÑANA'),
        "aula_id": postulante.aula_id,
        "activo": postulante.activo
    }


@router.post("/postulantes")
async def crear_postulante(
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Crea un nuevo postulante"""
    verificar_permisos_coordinador(usuario)
    
    # Verificar DNI único
    existe = db.query(Postulante).filter(Postulante.dni == data.get('dni')).first()
    if existe:
        raise HTTPException(status_code=400, detail="Ya existe un postulante con ese DNI")
    
    postulante = Postulante(
        dni=data.get('dni'),
        nombres=data.get('nombres', '').upper(),
        apellido_paterno=data.get('apellido_paterno', '').upper(),
        apellido_materno=data.get('apellido_materno', '').upper(),
        programa_educativo=data.get('programa_educativo'),
        turno=data.get('turno', 'MAÑANA'),
        proceso_admision=obtener_proceso_actual()
    )
    
    db.add(postulante)
    db.commit()
    db.refresh(postulante)
    
    return {"success": True, "id": postulante.id, "message": "Postulante registrado"}


@router.put("/postulantes/{postulante_id}")
async def actualizar_postulante(
    postulante_id: int,
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Actualiza un postulante"""
    verificar_permisos_coordinador(usuario)
    
    postulante = db.query(Postulante).filter(Postulante.id == postulante_id).first()
    if not postulante:
        raise HTTPException(status_code=404, detail="Postulante no encontrado")
    
    # Verificar DNI único si cambió
    if data.get('dni') and data.get('dni') != postulante.dni:
        existe = db.query(Postulante).filter(
            Postulante.dni == data.get('dni'),
            Postulante.id != postulante_id
        ).first()
        if existe:
            raise HTTPException(status_code=400, detail="Ya existe otro postulante con ese DNI")
    
    # Actualizar campos
    if data.get('dni'):
        postulante.dni = data.get('dni')
    if data.get('nombres'):
        postulante.nombres = data.get('nombres').upper()
    if data.get('apellido_paterno'):
        postulante.apellido_paterno = data.get('apellido_paterno').upper()
    if data.get('apellido_materno'):
        postulante.apellido_materno = data.get('apellido_materno').upper()
    if data.get('programa_educativo'):
        postulante.programa_educativo = data.get('programa_educativo')
    if data.get('turno'):
        postulante.turno = data.get('turno')
    
    db.commit()
    
    return {"success": True, "message": "Postulante actualizado"}


@router.delete("/postulantes/{postulante_id}")
async def eliminar_postulante(
    postulante_id: int,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Elimina un postulante (soft delete)"""
    verificar_permisos_coordinador(usuario)
    
    postulante = db.query(Postulante).filter(Postulante.id == postulante_id).first()
    if not postulante:
        raise HTTPException(status_code=404, detail="Postulante no encontrado")
    
    postulante.activo = False
    db.commit()
    
    return {"success": True, "message": "Postulante eliminado"}


@router.post("/postulantes/carga-masiva")
async def carga_masiva_postulantes(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Carga masiva de postulantes desde Excel o CSV"""
    verificar_permisos_coordinador(usuario)
    
    if not archivo.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Formato no soportado. Use Excel o CSV")
    
    proceso = obtener_proceso_actual()
    registrados = 0
    errores = []
    
    try:
        content = await archivo.read()
        
        if archivo.filename.endswith('.csv'):
            # Procesar CSV
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            # Procesar Excel
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    rows.append(dict(zip(headers, row)))
        
        for i, row in enumerate(rows, 2):
            try:
                dni = str(row.get('DNI', row.get('dni', ''))).strip()
                if not dni or len(dni) != 8:
                    errores.append(f"Fila {i}: DNI inválido")
                    continue
                
                # Verificar duplicado
                existe = db.query(Postulante).filter(Postulante.dni == dni).first()
                if existe:
                    errores.append(f"Fila {i}: DNI {dni} ya existe")
                    continue
                
                postulante = Postulante(
                    dni=dni,
                    nombres=str(row.get('Nombres', row.get('nombres', ''))).upper().strip(),
                    apellido_paterno=str(row.get('Apellido Paterno', row.get('apellido_paterno', ''))).upper().strip(),
                    apellido_materno=str(row.get('Apellido Materno', row.get('apellido_materno', ''))).upper().strip(),
                    programa_educativo=str(row.get('Programa', row.get('programa_educativo', ''))).strip(),
                    turno=str(row.get('Turno', row.get('turno', 'MAÑANA'))).upper().strip(),
                    proceso_admision=proceso
                )
                
                db.add(postulante)
                registrados += 1
                
            except Exception as e:
                errores.append(f"Fila {i}: {str(e)}")
        
        db.commit()
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")
    
    return {
        "success": True,
        "registrados": registrados,
        "errores": errores[:20] if errores else []  # Máximo 20 errores
    }


# ============================================================
# CRUD AULAS
# ============================================================

@router.get("/aulas")
async def listar_aulas(
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Lista todas las aulas"""
    proceso = obtener_proceso_actual()
    
    # Query con conteo de asignados
    result = db.execute(text("""
        SELECT 
            a.id, a.codigo, a.nombre, a.pabellon, a.piso, a.numero, 
            a.capacidad, a.activo,
            COUNT(ae.id) as asignados
        FROM aulas a
        LEFT JOIN asignaciones_examen ae ON ae.aula_id = a.id
        LEFT JOIN postulantes p ON p.id = ae.postulante_id AND p.proceso_admision = :proceso AND p.activo = true
        WHERE a.activo = true
        GROUP BY a.id, a.codigo, a.nombre, a.pabellon, a.piso, a.numero, a.capacidad, a.activo
        ORDER BY a.codigo
    """), {"proceso": proceso}).fetchall()
    
    aulas_data = []
    for a in result:
        aulas_data.append({
            "id": a.id,
            "codigo": a.codigo,
            "nombre": a.nombre,
            "programa": None,  # No existe este campo en el modelo
            "pabellon": a.pabellon,
            "piso": a.piso,
            "numero": a.numero,
            "capacidad": a.capacidad or 0,
            "asignados": a.asignados or 0,
            "activo": a.activo
        })
    
    return {"aulas": aulas_data}


@router.get("/aulas/siguiente-codigo")
async def siguiente_codigo_aula(
    prefijo: str = "AU",
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Obtiene el siguiente código de aula disponible"""
    
    # Buscar el último código con ese prefijo
    ultimo = db.query(Aula).filter(
        Aula.codigo.like(f"{prefijo}-%")
    ).order_by(Aula.codigo.desc()).first()
    
    if ultimo:
        try:
            ultimo_num = int(ultimo.codigo.split('-')[1])
            siguiente_num = ultimo_num + 1
        except:
            siguiente_num = 1
    else:
        siguiente_num = 1
    
    codigo = f"{prefijo}-{siguiente_num:03d}"
    
    return {"codigo": codigo, "prefijo": prefijo, "numero": siguiente_num}


@router.get("/aulas/{aula_id}")
async def obtener_aula(
    aula_id: int,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Obtiene un aula por ID"""
    aula = db.query(Aula).filter(Aula.id == aula_id).first()
    
    if not aula:
        raise HTTPException(status_code=404, detail="Aula no encontrada")
    
    return {
        "id": aula.id,
        "codigo": aula.codigo,
        "nombre": aula.nombre,
        "programa": getattr(aula, 'programa', None),
        "pabellon": aula.pabellon,
        "piso": aula.piso,
        "numero": aula.numero,
        "capacidad": aula.capacidad
    }


@router.post("/aulas")
async def crear_aula(
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Crea una nueva aula"""
    verificar_permisos_coordinador(usuario)
    
    # Generar código automático
    prefijo = data.get('prefijo', 'AU')
    resultado = await siguiente_codigo_aula(prefijo, db, usuario)
    codigo = resultado['codigo']
    
    # Verificar código único
    existe = db.query(Aula).filter(Aula.codigo == codigo).first()
    if existe:
        raise HTTPException(status_code=400, detail="El código de aula ya existe")
    
    aula = Aula(
        codigo=codigo,
        nombre=data.get('nombre'),
        pabellon=data.get('pabellon'),
        piso=data.get('piso'),
        numero=data.get('numero'),
        capacidad=int(data.get('capacidad', 30))
    )
    
    # Agregar programa si el campo existe
    if hasattr(aula, 'programa'):
        aula.programa = data.get('programa')
    if hasattr(aula, 'prefijo'):
        aula.prefijo = prefijo
    if hasattr(aula, 'proceso_admision'):
        aula.proceso_admision = obtener_proceso_actual()
    
    db.add(aula)
    db.commit()
    db.refresh(aula)
    
    return {"success": True, "id": aula.id, "codigo": codigo, "message": "Aula registrada"}


@router.put("/aulas/{aula_id}")
async def actualizar_aula(
    aula_id: int,
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Actualiza un aula"""
    verificar_permisos_coordinador(usuario)
    
    aula = db.query(Aula).filter(Aula.id == aula_id).first()
    if not aula:
        raise HTTPException(status_code=404, detail="Aula no encontrada")
    
    if data.get('nombre') is not None:
        aula.nombre = data.get('nombre')
    if data.get('pabellon') is not None:
        aula.pabellon = data.get('pabellon')
    if data.get('piso') is not None:
        aula.piso = data.get('piso')
    if data.get('numero') is not None:
        aula.numero = data.get('numero')
    if data.get('capacidad') is not None:
        aula.capacidad = int(data.get('capacidad'))
    if hasattr(aula, 'programa') and data.get('programa') is not None:
        aula.programa = data.get('programa')
    
    db.commit()
    
    return {"success": True, "message": "Aula actualizada"}


# ============================================================
# CRUD PROFESORES
# ============================================================

@router.get("/profesores")
async def listar_profesores(
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Lista todos los profesores"""
    
    profesores = db.query(Profesor).order_by(
        Profesor.apellido_paterno,
        Profesor.apellido_materno
    ).all()
    
    profesores_data = []
    for p in profesores:
        profesores_data.append({
            "id": p.id,
            "dni": p.dni,
            "nombres": p.nombres,
            "apellido_paterno": p.apellido_paterno,
            "apellido_materno": p.apellido_materno,
            "condicion": getattr(p, 'condicion', 'DOCENTE'),
            "email": p.email,
            "telefono": p.telefono,
            "habilitado": getattr(p, 'habilitado', True),
            "activo": p.activo
        })
    
    return {"profesores": profesores_data}


@router.get("/profesores/{profesor_id}")
async def obtener_profesor(
    profesor_id: int,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Obtiene un profesor por ID"""
    profesor = db.query(Profesor).filter(Profesor.id == profesor_id).first()
    
    if not profesor:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    
    return {
        "id": profesor.id,
        "dni": profesor.dni,
        "nombres": profesor.nombres,
        "apellido_paterno": profesor.apellido_paterno,
        "apellido_materno": profesor.apellido_materno,
        "condicion": getattr(profesor, 'condicion', 'DOCENTE'),
        "email": profesor.email,
        "telefono": profesor.telefono,
        "habilitado": getattr(profesor, 'habilitado', True)
    }


@router.post("/profesores")
async def crear_profesor(
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Crea un nuevo profesor"""
    verificar_permisos_coordinador(usuario)
    
    # Verificar DNI único
    existe = db.query(Profesor).filter(Profesor.dni == data.get('dni')).first()
    if existe:
        raise HTTPException(status_code=400, detail="Ya existe un profesor con ese DNI")
    
    profesor = Profesor(
        dni=data.get('dni'),
        nombres=data.get('nombres', '').upper(),
        apellido_paterno=data.get('apellido_paterno', '').upper(),
        apellido_materno=data.get('apellido_materno', '').upper(),
        email=data.get('email'),
        telefono=data.get('telefono')
    )
    
    # Campos opcionales si existen
    if hasattr(profesor, 'condicion'):
        profesor.condicion = data.get('condicion', 'DOCENTE')
    if hasattr(profesor, 'habilitado'):
        profesor.habilitado = True
    
    db.add(profesor)
    db.commit()
    db.refresh(profesor)
    
    return {"success": True, "id": profesor.id, "message": "Profesor registrado"}


@router.put("/profesores/{profesor_id}")
async def actualizar_profesor(
    profesor_id: int,
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Actualiza un profesor"""
    verificar_permisos_coordinador(usuario)
    
    profesor = db.query(Profesor).filter(Profesor.id == profesor_id).first()
    if not profesor:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    
    if data.get('dni'):
        profesor.dni = data.get('dni')
    if data.get('nombres'):
        profesor.nombres = data.get('nombres').upper()
    if data.get('apellido_paterno'):
        profesor.apellido_paterno = data.get('apellido_paterno').upper()
    if data.get('apellido_materno'):
        profesor.apellido_materno = data.get('apellido_materno').upper()
    if data.get('email') is not None:
        profesor.email = data.get('email')
    if data.get('telefono') is not None:
        profesor.telefono = data.get('telefono')
    if hasattr(profesor, 'condicion') and data.get('condicion'):
        profesor.condicion = data.get('condicion')
    
    db.commit()
    
    return {"success": True, "message": "Profesor actualizado"}


@router.put("/profesores/{profesor_id}/habilitacion")
async def toggle_habilitacion_profesor(
    profesor_id: int,
    data: dict,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Cambia el estado de habilitación de un profesor"""
    verificar_permisos_coordinador(usuario)
    
    profesor = db.query(Profesor).filter(Profesor.id == profesor_id).first()
    if not profesor:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    
    if hasattr(profesor, 'habilitado'):
        profesor.habilitado = data.get('habilitado', True)
    
    db.commit()
    
    return {"success": True, "message": "Estado actualizado"}


@router.post("/profesores/carga-masiva")
async def carga_masiva_profesores(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Carga masiva de profesores desde Excel o CSV"""
    verificar_permisos_coordinador(usuario)
    
    if not archivo.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Formato no soportado")
    
    registrados = 0
    errores = []
    
    try:
        content = await archivo.read()
        
        if archivo.filename.endswith('.csv'):
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    rows.append(dict(zip(headers, row)))
        
        for i, row in enumerate(rows, 2):
            try:
                dni = str(row.get('DNI', row.get('dni', ''))).strip()
                if not dni or len(dni) != 8:
                    errores.append(f"Fila {i}: DNI inválido")
                    continue
                
                existe = db.query(Profesor).filter(Profesor.dni == dni).first()
                if existe:
                    errores.append(f"Fila {i}: DNI {dni} ya existe")
                    continue
                
                profesor = Profesor(
                    dni=dni,
                    nombres=str(row.get('Nombres', row.get('nombres', ''))).upper().strip(),
                    apellido_paterno=str(row.get('Apellido Paterno', row.get('apellido_paterno', ''))).upper().strip(),
                    apellido_materno=str(row.get('Apellido Materno', row.get('apellido_materno', ''))).upper().strip(),
                    email=str(row.get('Email', row.get('email', ''))).strip() or None,
                    telefono=str(row.get('Telefono', row.get('telefono', ''))).strip() or None
                )
                
                if hasattr(profesor, 'condicion'):
                    condicion = str(row.get('Condicion', row.get('condicion', 'DOCENTE'))).upper().strip()
                    if condicion in ['DOCENTE', 'ADMINISTRATIVO', 'EXTERNO']:
                        profesor.condicion = condicion
                    else:
                        profesor.condicion = 'DOCENTE'
                
                db.add(profesor)
                registrados += 1
                
            except Exception as e:
                errores.append(f"Fila {i}: {str(e)}")
        
        db.commit()
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")
    
    return {
        "success": True,
        "registrados": registrados,
        "errores": errores[:20]
    }


# ============================================================
# LISTADOS
# ============================================================

@router.get("/listados/preview")
async def preview_listado(
    tipo: str = "alfabetico",
    aula_id: Optional[int] = None,
    programa: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Genera preview HTML de un listado"""
    proceso = obtener_proceso_actual()
    
    if tipo == "alfabetico":
        # Lista alfabética general
        rows = db.execute(text("""
            SELECT 
                p.id, p.dni, p.nombres, p.apellido_paterno, p.apellido_materno,
                p.programa_educativo,
                a.codigo as aula_codigo, a.pabellon, a.piso
            FROM postulantes p
            LEFT JOIN asignaciones_examen ae ON ae.postulante_id = p.id
            LEFT JOIN aulas a ON a.id = ae.aula_id
            WHERE p.proceso_admision = :proceso AND p.activo = true
            ORDER BY p.apellido_paterno, p.apellido_materno, p.nombres
        """), {"proceso": proceso}).fetchall()
        
        html = generar_html_listado_alfabetico(rows, proceso)
        
    elif tipo == "por-aula":
        if not aula_id:
            raise HTTPException(status_code=400, detail="Seleccione un aula")
        
        aula = db.execute(text("""
            SELECT id, codigo, nombre, pabellon, piso, capacidad
            FROM aulas WHERE id = :aula_id
        """), {"aula_id": aula_id}).fetchone()
        
        if not aula:
            raise HTTPException(status_code=404, detail="Aula no encontrada")
        
        rows = db.execute(text("""
            SELECT 
                p.id, p.dni, p.nombres, p.apellido_paterno, p.apellido_materno,
                p.programa_educativo
            FROM postulantes p
            JOIN asignaciones_examen ae ON ae.postulante_id = p.id
            WHERE ae.aula_id = :aula_id AND p.proceso_admision = :proceso AND p.activo = true
            ORDER BY p.apellido_paterno, p.apellido_materno
        """), {"aula_id": aula_id, "proceso": proceso}).fetchall()
        
        html = generar_html_listado_aula(aula, rows)
        
    elif tipo == "por-programa":
        rows = db.execute(text("""
            SELECT 
                p.id, p.dni, p.nombres, p.apellido_paterno, p.apellido_materno,
                p.programa_educativo,
                a.codigo as aula_codigo
            FROM postulantes p
            LEFT JOIN asignaciones_examen ae ON ae.postulante_id = p.id
            LEFT JOIN aulas a ON a.id = ae.aula_id
            WHERE p.proceso_admision = :proceso AND p.activo = true
            ORDER BY p.programa_educativo, p.apellido_paterno, p.apellido_materno
        """), {"proceso": proceso}).fetchall()
        
        html = generar_html_listado_programa(rows)
        
    else:
        raise HTTPException(status_code=400, detail="Tipo de listado no válido")
    
    return {"success": True, "html": html}


def generar_html_listado_alfabetico(rows, proceso):
    """Genera HTML para listado alfabético"""
    html = f"""
    <div class="listado-preview-header">
        <h2>LISTADO GENERAL DE POSTULANTES</h2>
        <p>Proceso de Admisión {proceso}</p>
        <p>Total: {len(rows)} postulantes</p>
    </div>
    <table class="table">
        <thead>
            <tr>
                <th>Nº</th>
                <th>Apellidos y Nombres</th>
                <th>DNI</th>
                <th>Programa</th>
                <th>Aula Asignada</th>
            </tr>
        </thead>
        <tbody>
    """
    
    for i, row in enumerate(rows, 1):
        aula_info = "-"
        if row.aula_codigo:
            aula_info = f"{row.aula_codigo}"
            if row.pabellon or row.piso:
                aula_info += f" ({row.pabellon or ''} Piso {row.piso or ''})"
        
        nombre = f"{row.apellido_paterno} {row.apellido_materno}, {row.nombres}"
        html += f"""
            <tr>
                <td>{i}</td>
                <td>{nombre}</td>
                <td>{row.dni}</td>
                <td>{row.programa_educativo or '-'}</td>
                <td>{aula_info}</td>
            </tr>
        """
    
    html += """
        </tbody>
    </table>
    <div class="footer" style="margin-top: 50px; text-align: center;">
        <p>_______________________________</p>
        <p>Firma y Sello</p>
    </div>
    """
    
    return html


def generar_html_listado_aula(aula, rows):
    """Genera HTML para listado por aula"""
    
    profesor_info = "Sin asignar"
    
    html = f"""
    <div class="listado-preview-header">
        <h2>LISTADO DE POSTULANTES - AULA {aula.codigo}</h2>
        <p>{aula.nombre or ''} - {aula.pabellon or ''} Piso {aula.piso or ''}</p>
        <p>Profesor Vigilante: {profesor_info}</p>
        <p>Total: {len(rows)} postulantes | Capacidad: {aula.capacidad or 0}</p>
    </div>
    <table class="table">
        <thead>
            <tr>
                <th>Nº</th>
                <th>DNI</th>
                <th>Apellidos y Nombres</th>
                <th>Programa</th>
                <th>Firma</th>
            </tr>
        </thead>
        <tbody>
    """
    
    for i, row in enumerate(rows, 1):
        nombre = f"{row.apellido_paterno} {row.apellido_materno}, {row.nombres}"
        html += f"""
            <tr>
                <td>{i}</td>
                <td>{row.dni}</td>
                <td>{nombre}</td>
                <td>{row.programa_educativo or '-'}</td>
                <td style="width: 100px;"></td>
            </tr>
        """
    
    html += """
        </tbody>
    </table>
    <div class="footer" style="margin-top: 50px;">
        <div style="display: flex; justify-content: space-between;">
            <div style="text-align: center;">
                <p>_______________________________</p>
                <p>Profesor Vigilante</p>
            </div>
            <div style="text-align: center;">
                <p>_______________________________</p>
                <p>Coordinador</p>
            </div>
        </div>
    </div>
    """
    
    return html


def generar_html_listado_programa(rows):
    """Genera HTML para listado por programa"""
    
    # Agrupar por programa
    programas = {}
    for row in rows:
        prog = row.programa_educativo or "Sin Programa"
        if prog not in programas:
            programas[prog] = []
        programas[prog].append(row)
    
    html = """
    <div class="listado-preview-header">
        <h2>LISTADO DE POSTULANTES POR PROGRAMA</h2>
    </div>
    """
    
    for programa, lista in programas.items():
        html += f"""
        <h3 style="margin-top: 20px; background: #f0f0f0; padding: 10px;">{programa} ({len(lista)} postulantes)</h3>
        <table class="table">
            <thead>
                <tr>
                    <th>Nº</th>
                    <th>DNI</th>
                    <th>Apellidos y Nombres</th>
                    <th>Aula</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for i, row in enumerate(lista, 1):
            nombre = f"{row.apellido_paterno} {row.apellido_materno}, {row.nombres}"
            aula_codigo = row.aula_codigo or "-"
            
            html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{row.dni}</td>
                    <td>{nombre}</td>
                    <td>{aula_codigo}</td>
                </tr>
            """
        
        html += "</tbody></table>"
    
    return html


# ============================================================
# PLANTILLAS
# ============================================================

@router.get("/plantillas/{tipo}")
async def descargar_plantilla(
    tipo: str,
    usuario: dict = Depends(obtener_usuario_actual)
):
    """Descarga plantilla CSV para carga masiva"""
    
    if tipo == "postulantes":
        contenido = "DNI,Nombres,Apellido Paterno,Apellido Materno,Programa,Turno\n"
        contenido += "12345678,JUAN CARLOS,GARCIA,LOPEZ,COMPUTACION E INFORMATICA,MAÑANA\n"
        filename = "plantilla_postulantes.csv"
        
    elif tipo == "profesores":
        contenido = "DNI,Nombres,Apellido Paterno,Apellido Materno,Condicion,Email,Telefono\n"
        contenido += "12345678,MARIA ELENA,RODRIGUEZ,SANCHEZ,DOCENTE,correo@ejemplo.com,999888777\n"
        filename = "plantilla_profesores.csv"
        
    else:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    
    return StreamingResponse(
        io.StringIO(contenido),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )